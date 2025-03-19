import json
from typing import List
import openpyxl
from openpyxl.drawing.image import Image
import cv2
import pandas as pd
import matplotlib.pyplot as plt
import os
import math

from data_preprocess.utils import ActionType


def read_json(rpath: str):
    """
    从指定路径读取JSON文件并返回解析后的数据
    
    参数:
        rpath: JSON文件的路径
    返回:
        解析后的JSON数据
    """
    with open(rpath, 'r', encoding='utf-8') as file:
        data = json.load(file)
        return data


def write_json(anns: List, wpath: str):
    """
    将数据列表写入JSON文件
    
    参数:
        anns: 要写入的数据列表
        wpath: 输出JSON文件的路径
    """
    json.dump(anns, open(wpath, "w"))


def add_visilize2screenshot(image_rpath, ann, tag):
    """
    在截图上添加可视化标记（红色圆圈）并保存为新图片
    
    参数:
        image_rpath: 原始图片路径
        ann: 包含动作信息的注释对象或字典
        tag: 新图片的标识标签
    
    返回:
        添加标记后的图片路径
    """
    if type(ann) == dict:
        if ann["action_type"] != "DUAL_POINT":
            return image_rpath

        touch_point, lift_point = ann["touch_point"], ann["lift_point"]
    else:
        if ann.action_type != ActionType.DualPoint:
            return image_rpath
        touch_point, lift_point = ann.touch_point, ann.lift_point

    # 计算点击点（触摸点和抬起点的中点）
    click_point = [(touch_point[0] + lift_point[0]) / 2, (touch_point[1] + lift_point[1]) / 2]

    # 读取图片并获取尺寸
    image = cv2.imread(image_rpath)
    height, width, _ = image.shape

    # 将相对坐标转换为绝对像素坐标
    x = int(click_point[0] * width)
    y = int(click_point[1] * height)

    # 在图片上绘制红色圆圈
    cv2.circle(image, (x, y), 20, (0, 0, 255), -1)

    # 生成新图片路径并保存
    image_wpath = image_rpath.split(".")[0] + f"_{tag}.png"
    cv2.imwrite(image_wpath, image) 

    return image_wpath.replace("\\", "/")


def write_to_excel(anns, wpath):
    """
    将注释数据写入Excel文件，包括图片和文本信息
    
    参数:
        anns: 注释数据列表
        wpath: 输出Excel文件的路径
    """
    wb = openpyxl.Workbook()
    ws = wb.active

    # 设置表头
    ws.cell(row=1, column=1, value="image")
    ws.cell(row=1, column=2, value="image(add point)")
    ws.cell(row=1, column=3, value="task")
    ws.cell(row=1, column=4, value="history action")
    ws.cell(row=1, column=5, value="current action")
    ws.cell(row=1, column=6, value="rating")
    ws.cell(row=1, column=7, value="explanation")

    # 填充数据
    for idx, ann in enumerate(anns, start=2):
        ws.cell(row=idx, column=3, value=ann["task"])
        ws.cell(row=idx, column=4, value="\n".join(ann["action_desc_list"]))
        ws.cell(row=idx, column=5, value=ann["action_desc_list"][ann["step_id"]])
        ws.cell(row=idx, column=6, value=ann["rating"])
        ws.cell(row=idx, column=7, value=ann["explanation"])

        # 添加原始图片
        img = Image(ann["image_list"][ann["step_id"]].replace("\\", "/"))
        img.width, img.height = (240, 480)
        ws.row_dimensions[idx].height = 400
        ws.add_image(img, f'A{idx}')
        
        # 添加带标记的图片
        img = Image(ann["add_point_image_list"][ann["step_id"]])
        img.width, img.height = (240, 480)
        ws.row_dimensions[idx].height = 400
        ws.add_image(img, f'B{idx}')

    # 调整列宽
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 20
    wb.save(wpath)


def parse_response(response):
    """
    尝试将响应文本解析为JSON对象
    
    参数:
        response: 包含JSON数据的响应文本
    
    返回:
        解析后的JSON对象，解析失败则返回-1
    """
    try:
        response = response.replace("```", "").replace("json", "")
        return json.loads(response)
    except:
        return -1
    

def write_jsonl(anns, wpath):
    """
    将数据列表写入JSONL文件（每行一个JSON对象）
    
    参数:
        anns: 数据列表
        wpath: 输出JSONL文件的路径
    """
    with open(wpath, 'w', encoding='utf - 8') as f:
        for item in anns:
            json_line = json.dumps(item)
            f.write(json_line + '\n')


def read_jsonl(rpath):
    """
    从JSONL文件读取数据
    
    参数:
        rpath: JSONL文件的路径
    
    返回:
        包含所有JSON对象的列表
    """
    data = []
    with open(rpath, 'r', encoding='utf - 8') as f:
        for idx, line in enumerate(f):
            line = line.strip()
            try:
                json_obj = json.loads(line)
                data.append(json_obj)
            except:
                print(f"Error decoding JSON on line: {idx}")
    return data


def read_xlsx(rpath):
    """
    读取Excel文件并转换为字典列表
    
    参数:
        rpath: Excel文件的路径
    
    返回:
        包含Excel数据的字典列表
    """
    data = pd.read_excel(rpath)
    return data.to_dict(orient="records")


def dict_mean(dict_list):
    """
    计算字典列表中各键值的平均值
    对于包含'min'的键取最小值，包含'max'的键取最大值
    
    参数:
        dict_list: 字典列表
    
    返回:
        包含平均值的字典
    """
    mean_dict = {}
    if len(dict_list) > 0:
        for key in dict_list[0].keys():
            if "min" in key:
                mean_dict[key] = min(d[key] for d in dict_list)
            elif "max" in key:
                mean_dict[key] = max(d[key] for d in dict_list)
            else:
                mean_dict[key] = sum(d[key] for d in dict_list) / len(dict_list)
    return mean_dict


def smooth(scalars: List[float]) -> List[float]:
    """
    使用指数加权平均对数据进行平滑处理
    
    参数:
        scalars: 需要平滑的数值列表
    
    返回:
        平滑后的数值列表
    """
    if len(scalars) == 0:
        return []

    last = scalars[0]
    smoothed = []
    # 使用sigmoid函数动态调整权重
    weight = 1.9 * (1 / (1 + math.exp(-0.05 * len(scalars))) - 0.5)
    for next_val in scalars:
        smoothed_val = last * weight + (1 - weight) * next_val
        smoothed.append(smoothed_val)
        last = smoothed_val
    return smoothed


def plot_loss(log_dir: str, keys: List[str] = ["loss"]) -> None:
    """
    绘制训练过程中的损失曲线并保存为图片
    
    参数:
        log_dir: 日志目录路径
        keys: 要绘制的指标名称列表，默认为["loss"]
    """
    plt.switch_backend("agg")  # 使用非交互式后端
    data = read_jsonl(os.path.join(log_dir, "train_log.jsonl"))

    for key in keys:
        steps, metrics = [], []
        for i in range(len(data)):
            if key in data[i]:
                steps.append(data[i]["step"])
                metrics.append(data[i][key])

        plt.figure()
        # 绘制原始数据和平滑后的数据
        plt.plot(steps, metrics, color="#1f77b4", alpha=0.4, label="original")
        plt.plot(steps, smooth(metrics), color="#1f77b4", label="smoothed")
        plt.title(f"{key} of {log_dir}")
        plt.xlabel("step")
        plt.ylabel(key)
        plt.legend()
        figure_path = os.path.join(log_dir, "training_{}.png".format(key))
        plt.savefig(figure_path, format="png", dpi=100)
        print("Figure saved at:", figure_path)
